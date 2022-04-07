#Import Modules
import praw
import requests
from bs4 import BeautifulSoup
import numpy as np
import pandas as pd
from praw.models import MoreComments 
from selenium import webdriver
import time
from datetime import date
import openpyxl
from openpyxl.styles import Font
import copy

#Scrape CoinMarketCap for daily price and top 10 coins via webdriver grab the data
coin_url = 'https://coinmarketcap.com/'
driver = webdriver.Chrome(chrome_driver_path)
driver.get(coin_url)
driver.execute_script("window.scrollTo(0, 1500)") #This scrolls down to allow ~20 coins to be loaded into coin_soup
coin_soup = BeautifulSoup(driver.page_source, 'html.parser')
driver.close()
coin_tag = coin_soup.find('div', class_ = 'h7vnx2-1 bFzXgL').find('tbody').find_all('tr') #All coins are in a <tr> tag

#Create top 10 within top_10_list as (Name), (Alias), (Price)
top_10_list = [(),(),(),(),(),(),(),(),(),()]
skip_list = ['usd coin', 'tether', 'bnb', 'binance usd', 'Crypto.com Coin']
coin_count = 0
for coin in coin_tag:
    if (coin.find_all('p')[1].text.lower()) not in skip_list:
        if coin_count <= 9:
            top_10_list[coin_count] = (coin.find_all('p')[1].text,coin.find_all('p')[2].text,coin.find_all('p')[3].text.split('$')[2]) 
            coin_count+=1
        else:
            break


#Top Coins by Marketcap
class Coin:
    #Create coin objects to track name and alias
    def __init__(self, name, alias):
        self.name = name
        self.alias = alias

#Define Coins
coin0 = Coin(top_10_list[0][0], top_10_list[0][1])
coin1 = Coin(top_10_list[1][0], top_10_list[1][1])
coin2 = Coin(top_10_list[2][0], top_10_list[2][1])
coin3 = Coin(top_10_list[3][0], top_10_list[3][1])
coin4 = Coin(top_10_list[4][0], top_10_list[4][1])
coin5 = Coin(top_10_list[5][0], top_10_list[5][1])
coin6 = Coin(top_10_list[6][0], top_10_list[6][1])
coin7 = Coin(top_10_list[7][0], top_10_list[7][1])
coin8 = Coin(top_10_list[8][0], top_10_list[8][1])
coin9 = Coin(top_10_list[9][0], top_10_list[9][1])
coin_list = [coin0, coin1, coin2, coin3, coin4, coin5, coin6, coin7, coin8, coin9]

#Load Date
today = date.today()
tab_date = today.strftime("%B %Y")
full_date = today.strftime("%B %d %Y")
dd_search_date = today.strftime("%B_%d").lower()

#Load database (Sheetnames are saved as: Month Year. tab_date is used to load the correct sheet.)
df_path = df_path
try:
    db_df = pd.read_excel(df_path, sheet_name = tab_date)
except:
    pass #You need to be in a catch for when a new month needs to be created


#Prepare Dataframe db_df for comment scrape
#Create Column for Any Coin not Previously in Top 10
for coin in coin_list:
    if coin.name.title() not in db_df.columns:
        db_df[coin.name.title()] = 0

#Drop Average and Total Columns
for col in db_df.columns:
    if 'Avg' in col or 'Total' in col:
        db_df.drop(columns = [col], inplace = True)

#Prepare day_dict for comment scrape/count
day_dict = {full_date: {}}
day_dict[full_date] = {
    coin0.name.title() : 0,
    coin1.name.title() : 0,
    coin2.name.title() : 0,
    coin3.name.title() : 0,
    coin4.name.title() : 0,
    coin5.name.title() : 0,
    coin6.name.title() : 0,
    coin7.name.title() : 0,
    coin8.name.title() : 0,
    coin9.name.title() : 0
}

#Parse Daily Discussion Links
#Grab daily discussion link based on the idea that Month_Date is included in the title.
parent_url = 'https://www.reddit.com/r/CryptoCurrency/search/?q=daily%20discussion&restrict_sr=1&sr_nsfw='
r = requests.get(parent_url)

#Parse Daily Discussion link
soup = BeautifulSoup(r.content, 'html.parser')
container = soup.find('div', class_='QBfRw7Rj8UkxybFpX-USO')
heads = container.find_all('a', href=True)

#Grab Daily Discussion
dd_link = ''
for head in heads:
    grab_href = head.get('href')
    if 'http' in grab_href and dd_search_date in grab_href:
        dd_link = grab_href

#Connect PRAW Agent
reddit_read_only = praw.Reddit(client_id = enter_your_id,
                    client_secret= enter_your_secret,
                    user_agent= enter_your_agent) #Obtained from reddit


#Scrape Comments from Daily Discussion Links
#Reddit utilizes JS to load more comments as you scroll down. Let's utilize PRAW to scrape comments
def grab_comments():
    submission = reddit_read_only.submission(url = dd_link)
    for coin in coin_list:
        count = 0
        for comment in submission.comments:
            if type(comment) == MoreComments:
                submission.comments.replace_more(limit = 0)
                for comment in submission.comments.list():
                    if coin.name.lower() in comment.body.lower() or coin.alias.lower() in comment.body.lower():
                        count+=1
            elif coin.name.lower() in comment.body.lower() or coin.alias.lower() in comment.body.lower():
                count+=1
        day_dict[full_date][(coin.name).title()] = count                

grab_comments()

#Append to Dataframe db_df
day_dict_df = pd.DataFrame(day_dict)

#Set index for db_df (.T transposes day_dict_df to match the formatting)
db_df = db_df.rename(
    columns = {'Unnamed: 0' :'Date'}).set_index('Date').append(day_dict_df.T).replace(np.nan, 0)

#Append to previous DataFrame
df_path = historical_dataframe_path
wb = openpyxl.load_workbook(df_path)
ws = wb[tab_date]
col_num = 0

ws_dict = {} #Temporary storage of ws column header and column number

#Create date
ws.cell(row = ws.max_row + 1, column = 1).value = full_date
ws.cell(row = ws.max_row + 1, column = 1).font = Font(bold=True)
max_rows = ws.max_row + 1

for i, cols in enumerate(ws.iter_cols(min_row = 1, max_row = 1, min_col = 2)):
    for col in cols:
        col_num = i + 2
        ws_dict[col.value] = col_num
for name in db_df.columns:
    if name in ws_dict: #Fill in new comment count for previously captured coin
        ws.cell(row = ws.max_row - 1, column = ws_dict[name]).value = db_df[name][full_date]
    elif name not in ws_dict: #Create new column
        print(name)
        ws.cell(row = 1, column = ws.max_column + 1).value = name
        ws.cell(row = ws.max_row + 1, column = ws.max_column + 1).value = db_df[name][full_date]
        #Fill empty cells in new column
        for rows in ws.iter_rows(min_row = 2, min_col = max_column, max_col = max_column):
            for row in rows:
                if row.value == None:
                    row.value = 'N/A'

#Append DataFrame.xlsx
wb.save(save_path)